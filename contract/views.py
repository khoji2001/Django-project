from __future__ import print_function
from mailmerge import MailMerge
from datetime import date,timedelta
from django.conf import settings
from supplier.models import supplier,coffer,issuer
from .models import select_MessageText
from mailmerge import MailMerge
from django.db.models import Sum
from config.settings import BASE_DIR
from zipfile import ZipFile
from os.path import basename
from document.models import  contract_document , customer_document , contract_directory_path ,customer_directory_path
from OTP.models import Emails
from django.shortcuts import render
from django.http import FileResponse, JsonResponse , HttpResponse, response
from django.contrib.auth.decorators import user_passes_test
from .models import  vcc, payment, contract  ,clear_receipt , add_card
from customer.models import customer
from supplier.models import supplier
from extensions import jalali
from extensions.utill import convert_to_englishnum,persian_numbers_converter
import os
from django.utils import timezone
from django.core.mail import EmailMessage
from .forms import *
from supplier.permissions import IsSupplier
import openpyxl
from rest_framework.status import HTTP_400_BAD_REQUEST, HTTP_200_OK, HTTP_302_FOUND, HTTP_404_NOT_FOUND, \
	HTTP_201_CREATED
from rest_framework.views import APIView
from .serializers import *
from rest_framework.permissions import IsAuthenticated 
from rest_framework.response import Response
from rest_framework.decorators import api_view , schema , permission_classes
from rest_framework.schemas import AutoSchema
from rest_framework import status
from OTP.views import send_sms

INPUT_DIR = os.path.join(settings.BASE_DIR,"contract_docs/input")
OUTPUT_DIR = os.path.join(settings.BASE_DIR,"contract_docs")
OUTPUTS_DIR = os.path.join(settings.BASE_DIR,"outputs")
# Create your views here.
def change_address(request):
	contracts = contract.objects.exclude(status__in=['5','6','7'])
	for c in contracts:
		text = """با سلام"""
		send_sms(c.customer_mobile(), text)
	return HttpResponse("موفق باشید")
@api_view()
@permission_classes([IsSupplier])
@schema(AutoSchema())
def c_receipt(request,c_id ,type):
    destination = os.path.join(OUTPUTS_DIR,"receipt")
    if not os.path.exists(destination):
        os.makedirs(destination)
    receipt = clear_receipt(c_id)
    try:
        doc_destination = os.path.join(destination,str(receipt.customer_mellicode)+'.docx')
        source = receipt.c_receipt_file_path
        docxMerge(vars(receipt),source,doc_destination,destination)
        pdf_destination = os.path.join(destination,str(receipt.customer_mellicode)+'.pdf')
        if (type == 'doc'):
            return FileResponse(open(doc_destination,'rb'),as_attachment = True)
        elif (type == 'pdf'):
            return FileResponse(open(pdf_destination,'rb'),as_attachment = True)
        else:
            return HttpResponse("404")
    except:
        return HttpResponse(" قرارداد مورد نظر یافت نشد یا تسویه نشده است.")

@user_passes_test(lambda u : u.is_superuser)
def final_c(request, c_id , type):
    destination = os.path.join(OUTPUTS_DIR,"contracts")
    if not os.path.exists(destination):
        os.makedirs(destination)
    contract = contract_kala(c_id)
    try:
        contract.vcc_number = persian_numbers_converter(contract.vcc_number)
        source = contract.file_path
        doc_destination = os.path.join(destination,str(contract.contract_id) + 'قرارداد.docx')
        docxMerge(vars(contract),source,doc_destination , destination)
        pdf_destination = os.path.join(destination,str(contract.contract_id) + 'قرارداد.pdf')
        if (type == 'doc'):
            return FileResponse(open(doc_destination,'rb'),as_attachment = True)
        elif (type == 'pdf'):
            if os.path.exists(pdf_destination):
                return FileResponse(open(pdf_destination,'rb'),as_attachment = True)
            else:
                return HttpResponse("اطلاعات متقاضی یا تامین کننده ناقص است")
        else:
            return HttpResponse("404")
    except Exception as e:
        return HttpResponse(str(e))

@api_view()
@permission_classes([IsSupplier])
@schema(AutoSchema())
def final_c_supp(request, c_id , type):
    user = request.user
    destination = os.path.join(OUTPUTS_DIR,"contracts")
    if not os.path.exists(destination):
        os.makedirs(destination)
    if user.supplier.Type == '1':
        contract = contract_kala(c_id)
        try:
            contract.vcc_number = persian_numbers_converter(contract.vcc_number)
            source = contract.file_path
            doc_destination = os.path.join(destination,str(contract.contract_id)+'قرارداد.docx')
            docxMerge(vars(contract),source,doc_destination , destination)
            pdf_destination = os.path.join(destination,str(contract.contract_id)+'قرارداد.pdf')
            if (type == 'doc'):
                return FileResponse(open(doc_destination,'rb'),as_attachment = True)
            elif (type == 'pdf'):
                if os.path.exists(pdf_destination):
                    return FileResponse(open(pdf_destination,'rb'),as_attachment = True)
                else:
                    return HttpResponse("اطلاعات متقاضی یا تامین کننده ناقص است")
            else:
                return HttpResponse("404")
        except Exception as e:
            return HttpResponse(str(e))
    else:
        return HttpResponse("403")

@user_passes_test(lambda u : u.is_superuser)
def coff_c(request, c_id , type):
    destination = os.path.join(OUTPUTS_DIR,"coff_contracts")
    if not os.path.exists(destination):
        os.makedirs(destination)
    contract = contract_kala(c_id)
    try:
        contract.vcc_number = convert_to_englishnum(contract.vcc_number , 'vcc_number')
        source = contract.coff_file_path
        doc_destination = os.path.join(destination,str(contract.contract_id)+'صندوق.docx')
        docxMerge(vars(contract),source,doc_destination , destination)
        pdf_destination = os.path.join(destination,str(contract.contract_id)+'صندوق.pdf')
        if (type == 'doc'):
            return FileResponse(open(doc_destination,'rb'),as_attachment = True)
        elif (type == 'pdf'):
            if os.path.exists(pdf_destination):
                return FileResponse(open(pdf_destination,'rb'),as_attachment = True)
            else:
                return HttpResponse("اطلاعات متقاضی یا تامین کننده ناقص است")
        else:
            return HttpResponse("404")
    except:
        return HttpResponse("قرارداد مورد نظر یافت نشد")

@api_view()
@permission_classes([IsSupplier])
@schema(AutoSchema())
def supp_c(request, c_id , type):
    user = request.user
    destination = os.path.join(OUTPUTS_DIR,"supp_contracts")
    if not os.path.exists(destination):
        os.makedirs(destination)
    if True:
        contract = contract_kala(c_id)
        try:
            contract.vcc_number = convert_to_englishnum(contract.vcc_number , 'vcc_number')
            source = contract.supp_file_path
            doc_destination = os.path.join(destination,str(contract.contract_id)+'فروشگاه.docx')
            docxMerge(vars(contract),source,doc_destination , destination)
            pdf_destination = os.path.join(destination,str(contract.contract_id)+'فروشگاه.pdf')
            if (type == 'doc'):
                return FileResponse(open(doc_destination,'rb'),as_attachment = True)
            elif (type == 'pdf'):
                if os.path.exists(pdf_destination):
                    return FileResponse(open(pdf_destination,'rb'),as_attachment = True)
                else:
                    return HttpResponse("اطلاعات متقاضی یا تامین کننده ناقص است")
            else:
                return HttpResponse("404")
        except:
            return HttpResponse("قرارداد مورد نظر یافت نشد")
    else:
        return HttpResponse("403")
    

@user_passes_test(lambda u : u.is_superuser)
def cust_c(request, c_id , type):
    destination = os.path.join(OUTPUTS_DIR,"cust_contracts")
    if not os.path.exists(destination):
        os.makedirs(destination)
    contract = contract_kala(c_id)
    try:
        contract.vcc_number = convert_to_englishnum(contract.vcc_number , 'vcc_number')
        source = contract.cust_file_path
        doc_destination = os.path.join(destination,str(contract.contract_id)+'مشتری.docx')
        docxMerge(vars(contract),source,doc_destination ,destination)
        pdf_destination = os.path.join(destination,str(contract.contract_id)+'مشتری.pdf')
        if (type == 'doc'):
            return FileResponse(open(doc_destination,'rb'),as_attachment = True)
        elif (type == 'pdf'):
            if os.path.exists(pdf_destination):
                return FileResponse(open(pdf_destination,'rb'),as_attachment = True)
            else:
                return HttpResponse("اطلاعات متقاضی یا تامین کننده ناقص است")
        else:
            return HttpResponse("404")
    except Exception as e:
        return HttpResponse(str(e))

@api_view()
@permission_classes([IsSupplier])
@schema(AutoSchema())
def cust_c_supp(request, c_id , type):
    user = request.user
    destination = os.path.join(OUTPUTS_DIR,"cust_contracts")
    if not os.path.exists(destination):
        os.makedirs(destination)
    if user.supplier.Type == '1':
        contract = contract_kala(c_id)
        try:
            contract.vcc_number = convert_to_englishnum(contract.vcc_number , 'vcc_number')
            source = contract.cust_file_path
            doc_destination = os.path.join(destination,f'{contract.contract_id}مشتری.docx')
            docxMerge(vars(contract),source,doc_destination , destination)
            pdf_destination = os.path.join(destination,f'{contract.contract_id}مشتری.pdf')
            if (type == 'doc'):
                return FileResponse(open(doc_destination,'rb'),as_attachment = True)
            elif (type == 'pdf'):
                if os.path.exists(pdf_destination):
                    return FileResponse(open(pdf_destination,'rb'),as_attachment = True)
                else:
                    return HttpResponse("اطلاعات متقاضی یا تامین کننده ناقص است")
            else:
                return HttpResponse("404")
        except Exception as e:
            return HttpResponse(str(e))
    else:
        return HttpResponse("403")
@api_view(['POST'])
@permission_classes([IsSupplier])
@schema(AutoSchema())
def receive_downpayment(request):
    data = dict(request.data.items())
    if data.get("contractID") is None:
        return Response({"Error" : "شناسه قرارداد ضروری است"} , status.HTTP_400_BAD_REQUEST)

    try:
        c = contract.objects.get(pk = data.get("contractID"))
        if int(c.status) == 0:
            c.status = '1'
            c.save()
            return Response({'message' : 'تایید پیش پرداخت ثبت شد'} , status.HTTP_200_OK)
        else:
            return Response({"Error" : ' عملیات امکان پذیر نیست '} , status.HTTP_405_METHOD_NOT_ALLOWED)
    except Exception as e:
        return Response({"Error" : str(e)} , status.HTTP_404_NOT_FOUND)

@api_view(['POST'])
@permission_classes([IsSupplier])
@schema(AutoSchema())
def receive_contract(request):
    data = dict(request.data.items())
    try:
        contrctsss = data['contr']
        Serializer = contractobjSerializer(data=contrctsss)
        file_path = os.path.join(INPUT_DIR,'selected_contracts_excel.xlsx')
        ex = openpyxl.load_workbook(file_path)
        sheet = ex['Sheet1']
        x = 1
        for item in Serializer.show(data=contrctsss):
            x += 1
            c = item
            sheet.cell(row=x, column=2).value = c.customer.full_name()
            sheet.cell(row=x, column=1).value = x - 1
            sheet.cell(row=x, column=3).value = c.net_amount
            sheet.cell(row=x, column=4).value = c.downpayment
            sheet.cell(row=x, column=5).value = c.supplier_balance
            sheet.cell(row=x, column=6).value = c.jinvoice_date
            sheet.cell(row=x, column=7).value = c.jsign_date()
            sheet.cell(row=x, column=8).value = c.jclear_date
            sheet.cell(row=x, column=9).value = c.customer.get_status_display()

            ex.save(file_path)
        return FileResponse(open(file_path, 'rb'), as_attachment=True)
    except Exception as e:
        return Response({"Error" : str(e)} , status.HTTP_404_NOT_FOUND)


def docxMerge(context,source,destination,pdf_dest): #mailmerge source docx template and generate contract
    template = source
    document = MailMerge(template)
    document.merge(**context)
    document.write(destination)
    document.close()
    os.system ("libreoffice --headless --convert-to pdf --outdir "+pdf_dest + " "+ destination)

@user_passes_test(lambda u : u.is_superuser)
def vccs(request):
    if request.method == 'POST':
        vcc_form = Uploadvccexcel(request.POST , request.FILES)
        vccs = {}
        if vcc_form.is_valid():
            vcc_excel = request.FILES['file']
            wb = openpyxl.load_workbook(vcc_excel)
            worksheet = wb[wb.sheetnames[0]]
            numbers = worksheet['A']
            for i in range(len(numbers)):
                if (numbers[i].value != None):
                    vccs[i] = numbers[i].value
                    add_card(numbers[i].value)

        return JsonResponse({'vccs' :vccs,})
    else:
        vcc_form = Uploadvccexcel()
        return render(request, 'vcc_upload.html', { 'vcc_form' : vcc_form , })

def notify_debt_contracts():
    debts = {}
    debt_contracts = contract.debt_contracts.all()
    for c in debt_contracts:
        text = select_MessageText().notify_debit.format(
    c.instalment_due_day , c.vcc_number ,c.old_debit_amount_persian)
        send_sms(c.customer_mobile(), text)

class detailpdf:
    def __init__(self,cc):
        self.vcc_number = cc.vcc_number
        self.contract_id = cc.contract_id
        try:
            self.customer_name = f'{cc.customer.first_name} {cc.customer.last_name}'
        except:
            self.customer_name = ''
        try:
            self.customer_mellicode = cc.customer.melli_code
        except:
            self.customer_mellicode = ''
        self.instalment_amount = cc.instalment_amount
        self.start_date = cc.j_instalment_start_date()

def notify_clear_contracts():
    today = persian_numbers_converter(jalali.Gregorian(timezone.now().date()).persian_string("{}/{}/{}"))
    clear_contracts = contract.clear_contracts.all()
    for c in clear_contracts:
        try:
            x = c.vcc.new_status
        except:
            x = ''
        if c.instalment_detail[0] <= 0 or x == '1' :
            try:
                balance_date = persian_numbers_converter(jalali.Gregorian(c.last_pay_date).persian_string("{}/{}/{}"))
            except:
                balance_date = ''
            y = detailpdf(c)
            document = MailMerge(Emails.objects.get(email_type = '3').file.path)
            context = {"date" : str(today) ,
              "contract_id" :  str(y.contract_id),
              "customer_name" : str(y.customer_name), 
              "customer_mellicode" : str(persian_numbers_converter(y.customer_mellicode)),
              "vcc_number" : str(y.vcc_number),
              "instalment_amount" : str(persian_numbers_converter(y.instalment_amount , 'price')),
              "start_date" : str(y.start_date),
              "clear_date" : str(balance_date),
            }
            document.merge(**context)
            destination = os.path.join(OUTPUT_DIR, 'tass',f'تسویه{y.contract_id}.docx')
            document.write(destination)
            document.close()

            
            email_dict = {
                'c_name' : y.customer_name,
                'today' : today,
                'contract_id' : y.contract_id,
                }
            
            emaill = Emails.objects.get(email_type = '3')
            too = [x.strip() for x in emaill.TO.split(',')]
            ccc = [x.strip() for x in emaill.cc.split(',')]
            
            email = EmailMessage(subject = emaill.ST.format(**email_dict),body = emaill.ET ,from_email = 'admin@test.com', to = too , cc = ccc ,headers= {'Content_Type' :'text/plain'})
            
            pdf_dest = os.path.join(OUTPUT_DIR, "tass" )
            os.system ("libreoffice --headless --convert-to pdf --outdir "+pdf_dest + " "+ destination)
            pdf_dest = os.path.join(OUTPUT_DIR, "tass",f'تسویه{y.contract_id}.pdf')
            email.attach_file(pdf_dest)
            email.send()
            os.remove(pdf_dest)
            os.remove(destination)
            c.status = '5'
            c.save()

        elif  x in ['0','2']: 
            if x == '0':
                email_dict = {
                'c_name' : c.customer.full_name(),
                'mellicode' : c.customer_mellicode,
                'instalment_detail' : c.instalment_detail[0],
                }
                emaill = Emails.objects.get(email_type = '6')
                too = [x.strip() for x in emaill.TO.split(',')]
                ccc = [x.strip() for x in emaill.cc.split(',')]
                email = EmailMessage(subject = emaill.ST,body = emaill.ET.format(**email_dict) ,from_email = 'admin@test.com', to = too , cc = ccc ,headers= {'Content_Type' :'text/plain'})
                email.send()
            if x == '2':
                sms_text = select_MessageText().mulct_pay.format(c.customer.full_name(),c.instalment_detail[0])
                send_sms(c.customer.user.mobile_number, sms_text)
       

        
@user_passes_test(lambda u : u.is_superuser)
def payments(request):
    if request.method == 'POST':
        payment_form = Uploadpaymentexcel(request.POST, request.FILES)
        if payment_form.is_valid():
            payment_excel = request.FILES['file']
            wb = openpyxl.load_workbook(payment_excel)
            worksheet = wb[wb.sheetnames[0]]
            dates = worksheet['D']
            vccs = worksheet['G']
            amounts = worksheet['K']
            voucher_ids = worksheet['F']
            payments = {}
            for i in range(len(dates)):
                if (dates[i].value != None and dates[i].value != 'تاریخ'):
                    vcc_fields = vccs[i].value.split('_')
                    if (len(vcc_fields) > 1):
                        payments[i] = (vcc_fields[1],dates[i].value,
                        amounts[i].value,voucher_ids[i].value,
                        payment.add_payment(vcc_fields[1],dates[i].value,
                        amounts[i].value,voucher_ids[i].value))
            if request.POST.get('sms_send') == 'True':
                notify_debt_contracts()
            notify_clear_contracts()
            return JsonResponse({'payments' :payments})
    else:
        payment_form = Uploadpaymentexcel()
        return render(request, 'payment_upload.html', { 'payment_form' : payment_form , })

SUPPLIERS = {
	2 : "تست" ,
	3 : "دیجی کالا " ,
	7 : "اسنپ",
}
@user_passes_test(lambda u : u.is_superuser)
def addcontracts(request):
    if request.method == 'POST':
        cont_form = Uploadcontexcel(request.POST, request.FILES)
        if cont_form.is_valid():
            cont_excel = request.FILES['file']
            wb = openpyxl.load_workbook(cont_excel)
            for val in SUPPLIERS.values():
                worksheet = wb[val]
                s = supplier.objects.get(name = val)
                cust_mobiles = worksheet['G']
                vccs = worksheet['B']
                contract_ids = worksheet['D']
                start_dates_year = worksheet['L']
                start_dates_month = worksheet['K']
                start_dates_day = worksheet['J']
                net_amounts = worksheet['M']
                dowpayments = worksheet['S']
                discounts = worksheet['V']
                number_of_instalments = worksheet['R']
                company_gains = worksheet['W']
                for i in range(1,len(contract_ids)):
                    if (contract_ids[i].value != None):
                        try:
                            c = customer.objects.get(user__mobile_number = str(cust_mobiles[i].value))
                        except:
                            pass
                        try:
                            v = vcc.objects.get(number = str(vccs[i].value))
                        except:
                            v = vcc.objects.create(number = str(vccs[i].value))
                        company_gain_rate = company_gains[i].value / net_amounts[i].value
                        year = int('13' + str(start_dates_year[i].value))
                        month = int(start_dates_month[i].value)
                        day = int(start_dates_day[i].value)

                        if month == 1:
                            year -= 1
                        month = ((month - 1) - 1 ) % 12 + 1
                        temp_date = jalali.Persian(year,month,day).gregorian_datetime()
                        sign_date = temp_date - timedelta(days =7)
                        try:
                            c = contract.objects.create(supplier = s,customer = c ,net_amount = net_amounts[i].value,
                            sign_date = sign_date, status = '4' , downpayment = dowpayments[i].value , 
                            discount= discounts[i].value , company_gain_rate = company_gain_rate,
                            contract_id = contract_ids[i].value ,number_of_instalment = number_of_instalments[i].value)
                            if c.number_of_instalment == 6:
                                c.status = '5'
                            c.Type = '0'
                            c.vcc = v
                            v.amount = 0
                            c.save()
                            v.save()
                        except:
                            pass

            return HttpResponse(type + 'با موفقیت اضافه شد')
    else:
        cont_form = Uploadcontexcel()
        return render(request, 'cont_upload.html', { 'cont_form' : cont_form , })


class ContractView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        try:
            c = user.customer
            queryset = c.contract_set.all().order_by("sign_date")
            serializer = ContractCustSerializer(queryset , many = True)
            
            return Response(serializer.data)
        except Exception as e:
            return Response({"Error" : str(e)} , status.HTTP_404_NOT_FOUND)

@api_view()
@permission_classes([IsAuthenticated])
@schema(AutoSchema())
def contractPayments(request,pk):
    user = request.user
    try:
        c = user.customer
        queryset = c.contract_set.get(pk =pk)
        serializer = ContractPaymentsSerializer(queryset)
        return Response(serializer.data)
    except Exception as e:
        try:
            s = user.supplier
            queryset = s.contract_set.get(pk =pk)
            serializer = ContractPaymentsSerializer(queryset)
            return Response(serializer.data)
        except Exception as w:
            return Response({"error":str(e) + str(w)} , status.HTTP_404_NOT_FOUND)
@api_view()
@permission_classes([IsAuthenticated])
@schema(AutoSchema())
def contractInstalments(request,pk):
    user = request.user
    try:
        c = user.customer
        queryset = c.contract_set.get(pk =pk)
        serializer = ContractInstalmentsSerializer(queryset)
        return Response(serializer.data)
    except Exception as e:
        try:
            s = user.supplier
            queryset = s.contract_set.get(pk =pk)
            serializer = ContractInstalmentsSerializer(queryset)
            return Response(serializer.data)
        except Exception as w:
            return Response({"error":str(e) + str(w)} , status.HTTP_404_NOT_FOUND)
@api_view()
@permission_classes([IsAuthenticated])
@schema(AutoSchema())
def contractDetail(request,pk):
    user = request.user
    try:
        c = user.customer
        queryset = c.contract_set.get(pk =pk)
        serializer = ContractCustSerializer(queryset)
        return Response(serializer.data)
    except Exception as e:
        try:
            s = user.supplier
            queryset = s.contract_set.get(pk =pk)
            serializer = ContractCustSerializer(queryset , context = {'supplier_type' : s.Type})
            return Response(serializer.data)
        except Exception as w:
            return Response({"error":str(e) + str(w)} , status.HTTP_404_NOT_FOUND)

@user_passes_test(lambda u : u.is_superuser)
def contractAdminDetail(request,pk):
    try:
        queryset = contract.objects.get(pk =pk)
        serializer = ContractCustSerializer(queryset)
        return JsonResponse(serializer.data)
    except Exception as e:
        return JsonResponse({"error":str(e)})

import jdatetime
PRE_GENDER = {
	'm' : 'حضرتعالی' , 
	'f' : 'سرکارعالی'
}
@user_passes_test(lambda u : u.is_superuser)
def due_contracts(request):
    if request.method == 'POST':
        due_form = DueContracts(request.POST)
        if due_form.is_valid():
            due_date = due_form.cleaned_data["due_date"]
            if due_date < date.today():
                flag = True
            else:
                flag = False
            due_contracts = contract.get_dueconts(due_date)
            due_date = persian_numbers_converter(jdatetime.date.fromgregorian(date = due_date + timedelta(days = 1),locale = "fa_IR").strftime("%a %d %b"))
            count = len(due_contracts)
            result = ""
            for c in due_contracts:
                try:
                    if c.number_of_pays() == 0:
                        is_first = "اول "
                    else:
                        is_first = ''

                    if c.status in ['2','3']:
                        c.status = '4'
                        c.save()
                    if flag:
                        instalment_amount = 0
                    else:
                        instalment_amount = c.instalment_amount
                    if instalment_amount + c.debit_amount() > 0:
                        amount = persian_numbers_converter(instalment_amount + c.debit_amount(),'price')
                        text = select_MessageText().notify_instalment.format(is_first,due_date,amount,c.vcc_number)
                        send_sms(c.customer.user.mobile_number , text)
                    else:
                        text = select_MessageText().good_payer.format(due_date , PRE_GENDER[c.customer.user.gender])
                        send_sms(c.customer.user.mobile_number , text)
                except Exception as e:
                    result += str(e) + '\n'
            send_sms("09374788634" , "پیامک یادآوری اقساط برای {0}نفر ارسال شد".format(str(count)))
            return HttpResponse(result)
    else:
        due_form = DueContracts()
        return render(request, 'due_contracts.html', { 'due_form' : due_form , })
def all_doc_contr(request,c_id):
    contr = contract.objects.get(pk = c_id)
    cust = customer.objects.get(contract = contr)
    suretyy = ContractSuretys.objects.filter(cont=contr)

    has_customer_info = 0
    has_surety_info = 0
    has_contract_info = 0

    contract_docs = contract_document.objects.filter(contract=contr)

    if not os.path.exists("contract_doc/"):
        os.makedirs("contract_doc")
    all_doc_destination = "contract_doc/"+str(contr.contract_id)+'.zip'
    zip_doc = ZipFile(all_doc_destination, 'w')

    with ZipFile('customer.zip', 'w') as zipObj:
        try:
            for folderName, subfolders, filenames in os.walk(os.path.join(BASE_DIR ,"media/")+customer_directory_path(customer_document.objects.get(customer= cust))):
                for filename in filenames:
                    has_customer_info = 1
                    filePath = os.path.join(folderName, filename)
                    zipObj.write(filePath, basename(filePath))
        except:
            pass
    if has_customer_info == 1:
        zip_doc.write('customer.zip')
    os.remove('customer.zip')

    if len(suretyy) > 0:	
        for index,sur in enumerate(suretyy):
            name_of_zip_file ='surety_doc'+ str(index+1) +'.zip' 
            with ZipFile(name_of_zip_file, 'w') as zipObj:
                customer_id = sur.surt
                try:
                    for folderName, subfolders, filenames in os.walk(os.path.join(BASE_DIR ,"media/")+customer_directory_path(customer_document.objects.get(customer=customer_id))):
                        for filename in filenames:
                            has_surety_info = 1
                            filePath = os.path.join(folderName, filename)
                            zipObj.write(filePath, basename(filePath))
                except:
                    pass
            if has_surety_info == 1:
                zip_doc.write(name_of_zip_file)
            has_surety_info = 0
            os.remove(name_of_zip_file)	

    if True :
        for index, contract_doc in enumerate(contract_docs):
            name_of_zip_file = 'contract_doc'+ str(index+1)  +'.zip' 
            with ZipFile(name_of_zip_file, 'w') as zipObj:
                try:
                    for folderName, subfolders, filenames in os.walk(os.path.join(BASE_DIR ,"media/")+contract_directory_path(contract_doc)):
                        for filename in filenames:
                            has_contract_info = 1
                            filePath = os.path.join(folderName, filename)
                            zipObj.write(filePath, basename(filePath))
                except:
                    pass
            if has_contract_info == 1:
                zip_doc.write(name_of_zip_file)
            has_contract_info = 0
            os.remove(name_of_zip_file)	
    zip_doc.close()
    return FileResponse(open(all_doc_destination,'rb'),as_attachment = True)

def n(field):
    return len(field)
def p(num):
    return persian_numbers_converter(num,'price')

@user_passes_test(lambda u : u.is_superuser)
def financing(request):
    if request.method == 'POST':
        financing_form = FinancingForm(request.POST)
        if financing_form.is_valid():
            btn_coff = financing_form.cleaned_data['btnـcoffer']
            btn_iss = financing_form.cleaned_data['btn_issuer']
            if btn_coff == True and btn_iss == True:
                iss = issuer.objects.all()
                coff = coffer.objects.all()
            elif btn_coff == True and btn_iss == False:
                coff = coffer.objects.all()
                iss = financing_form.cleaned_data["issuer"]
            elif btn_coff == False and btn_iss == True:
                coff = financing_form.cleaned_data["coffer"]
                iss = issuer.objects.all()
            else:
                iss = financing_form.cleaned_data["issuer"]
                coff = financing_form.cleaned_data["coffer"]
            
            coff_id =[]
            for cof in coff:
                coff_id.append(cof.id)
            iss_id =[]
            for isu in iss:
                iss_id.append(isu.id)
            all = contract.objects.filter(coffer__in = coff, issuer__in = iss)
            l1 = n(all)
            all_cancel = len(all.filter(status = '7'))
            l2 = all_cancel
            process = len(all) - all_cancel
            all_final = all.exclude(status__in = ['0','1','7'])
            tot_pay = all_final.aggregate(tot_pay = Sum('face_net_amount'))
            oo = all.filter(status = '5')
            ta = all.filter(status = '5').aggregate(tot_pay = Sum('face_net_amount'))
            ty = all.filter(status__in = ['0','1'])
            x = 0
            for con in ty:
                x += con.loan_amount 
            l7 = x
            y =0
            for con in all_final:
                y += con.loan_amount 
            
            l8 = y
            q = 0
            for con in all_final:
                q += con.total_amount_of_instalments
            l9 = q
            rr = all.exclude(status__in = ['5','7'])
            l10 = n(rr)
            ta = rr.aggregate(tot_pay = Sum('face_net_amount'))
            l11 = ta['tot_pay']
            w = 0
            for con in rr:
                w += con.loan_amount 
            l12 = w
            
            k = 0
            for con in rr:
                k += con.total_amount_of_instalments
            l177 = k
            all_coff = contract.objects.filter(supplier__coffer__pk__in = coff_id).annotate(pay = Sum('payments__amount')).aggregate(tot_pay = Sum('pay'))
            xx = 0
            for c in all_final:
                year_3,month_3,day_3 = jalali.Gregorian(date.today()).persian_tuple()
                years,months,dayss= jalali.Gregorian(c.instalment_start_date).persian_tuple()
                num_3 = (year_3 - years)*12 + month_3 - months
                if day_3 > dayss:
                    num_3 += 1
                xx += (num_3*c.instalment_amount)
            st = []
            for item in iss:
                st.append(item.name)
            sq = []
            for item in coff:
                sq.append(item.name)

            document = MailMerge(os.path.join(INPUT_DIR,"financing_mailmerge.docx"))
            context = {"date" : str(persian_numbers_converter(jalali.Gregorian(date.today()).persian_string('{}/{}/{}'))) ,
             "issuer_name" : ','.join(st) ,
             "coffer_name" : ','.join(sq) ,
              "tashkil_count" : str(p(l1)) , 
              "enseraf_count" : str(p(l2)),
              "nahai_count" : str(p(n(all_final))),
              "nahai_net_amount" : str(p(tot_pay['tot_pay'])),
              "tasvieh_count" : str(n(oo)),
              "tasvieh_net_amount" : str(p(ta['tot_pay'])),
              "entezar_loan_amount" : str(p(l7)),
              "variz_loan_amount" : str(p(l8)),
              "total_insts" : str(p(l9)),
              "due_insts" : str(p(xx)),
              "active_count" : str(p(l10)),
              "active_net_amount" : str(p(ta['tot_pay'])) ,
              "active_loan_amount" : str(p(l12)) ,
              "active_total_insts" : str(p(l177)),
              "all_pays" : str(p(all_coff['tot_pay'])),
            }
            document.merge(**context)
            destination = os.path.join(INPUT_DIR,"financing_mailmerge_out.docx")
            document.write(destination)
            document.close()
            pdf_dest = os.path.join(OUTPUT_DIR,"financing")
            os.system ("libreoffice --headless --convert-to pdf --outdir "+pdf_dest + " "+ destination)
            pdf = os.path.join(pdf_dest,"financing_mailmerge_out.pdf")
            os.remove(destination)
            return FileResponse( open(pdf,'rb'),as_attachment = True)
    else:
        finance_form = FinancingForm()
        return render(request, 'FinancingForm.html', { 'FinancingForm' : finance_form , })

@user_passes_test(lambda u : u.is_superuser)
def vcc_btn(request):
    if request.method == 'POST':
        financing_form = vccForm(request.POST)
        if financing_form.is_valid():
            btn_status = financing_form.cleaned_data['btn']
            if btn_status == True:
                supp = supplier.objects.all()
            else:
                supp = financing_form.cleaned_data["supp"]

            all = contract.objects.filter(supplier__in = supp)
            rr = all.exclude(status__in = ['5','7'])
            l1 = n(rr)

            tot_pay = rr.aggregate(tot_pay = Sum('face_net_amount'))
            l2 = tot_pay['tot_pay']

            all_final = all.exclude(status ='7').annotate(pay = Sum('payments__amount'))
            x = 0
            for c in all_final:
                if not isinstance(c.instalment_start_date, str):
                    year_3,month_3,day_3 = jalali.Gregorian(date.today()).persian_tuple()
                    years,months,dayss= jalali.Gregorian(c.instalment_start_date).persian_tuple()
                    num_3 = (year_3 - years)*12 + month_3 - months
                    if day_3 > dayss:
                        num_3 += 1
                    x += (num_3*c.instalment_amount)
            l3 = x
            if l3 == 0:
                l3 = 1
            all_fina = all.exclude(status = '7').annotate(pay = Sum('payments__amount')).aggregate(tot_pay = Sum('pay'))
            l4 = all_fina['tot_pay']
            if l4 is None:
                l4 = 1
            pa = all.filter(status = '4')
            z = 0
            aa = 0
            sa = 0
            saa = 0
            for con in pa:
                po = con.debit_amount()
                if po > 0:
                    qq = po // con.instalment_amount
                    
                    if qq < (con.number_of_instalment// 12) +1:
                        z += 1
                        aa += po
                    else:
                        sa += 1
                        saa += po
            l6 = z
            l7 = aa
            l8 = sa
            l9 = saa 
            per1 = round(aa/l3,3)*100
            q = saa/l3*100
            per2 = round(q,2)
            nok = all.filter(status = '6')
            ee = 0
            for c in nok:
                ee += c.debit_amount()
            
            www = 0
            for c in nok:
                www += c.total_amount_of_instalments
            
            l10 = z + sa +n(nok)
            l11 = ee + saa + aa
            nokk = nok.annotate(pay = Sum('payments__amount')).aggregate(tot_pay = Sum('pay'))
            payy = nokk['tot_pay']
            sxs = www/l3*100
            per3 = round( sxs, 2)
            vcc_tot  = z + sa + n(nok)
            vcc_tot_debit = aa + saa + ee
            per4 = round(vcc_tot_debit/l3,3)*100
            if payy is None:
                payy = 0
            nok1 = www - payy 
            vcc_q = nok1 + aa + saa
            per5 = round(vcc_q/l3,3)*100
            st = []
            for item in supp:
                st.append(item.name)
            sq = []
            for item in supp:
                sq.append(item.coffer.name)
            sq  = set(sq)
            ff = (l11/l3)*100
            per8 = round( ff , 2)
            document = MailMerge(os.path.join(INPUT_DIR,"vcc_btn_mailmerge.docx"))
            
            context = {"date" : str(persian_numbers_converter(jalali.Gregorian(date.today()).persian_string('{}/{}/{}'))) ,
            "supplier_name" : ','.join(st),
            "coffer_name" :  ','.join(sq),
            "active_count" : str(p(l1)), 
            "active_net_amount" : str(p(l2)),
            "due_insts" : str(p(l3)),
            "all_pays" : str(p(l4)),
            "coffer_beg" : str(p(l3 - l4)) ,
            "takhir_count" : str(p(l6)),
            "takhir_debt" : str(p(l7)),
            "takhir_rate" : str(p(per1)),
            "tavigh_count" : str(p(l8)),
            "tavigh_debt" : str(p(l9)),
            "tavigh_rate" : str(p(per2)),
            "nokol_rate" : str(p(round(ee/l3 , 3)*100)),
            "nokol_count" : str(p(n(nok))),
            "nokol_debt" : str(p(ee) ),
            "nokol_totalins" : str(p(www) ),
            "nokol_totalins_rate" : str(p(per3)),
            "vcc_count" : str(p(l10)),
            "vcc_debt" : str(p(l11)),
            "vcc_rate" : str(p(per8)),
            "vcc_total_debt" : str(p(vcc_q)),
            "vcc_total_rate" : str(p(round(vcc_q/l3,3)*100))
            }
            document.merge(**context)
            destination = os.path.join(INPUT_DIR,"vcc_btn_mailmerge_out.docx")
            document.write(destination)
            document.close()
            pdf_dest = os.path.join(OUTPUT_DIR,"vcc_btn")
            os.system ("libreoffice --headless --convert-to pdf --outdir "+pdf_dest + " "+ destination)
            pdf = os.path.join(pdf_dest,"vcc_btn_mailmerge_out.pdf")
            return FileResponse(open(pdf,'rb'),as_attachment = True)
    else:
        finance_form = vccForm()
        return render(request, 'vccForm.html', { 'vccForm' : finance_form , })
@user_passes_test(lambda u : u.is_superuser)
def supp(request):
    if request.method == 'POST':
        financing_form = vccForm(request.POST)
        if financing_form.is_valid():

            btn_status = financing_form.cleaned_data['btn']
            if btn_status == True:
                supp = supplier.objects.all()
            else:
                supp = financing_form.cleaned_data["supp"]
            all = contract.objects.filter(supplier__in = supp )
            tot_pay = all.aggregate(tot_pay = Sum('face_net_amount'))
            tot_payments = all.annotate(pay = Sum('payments__amount')).aggregate(tot_pay = Sum('pay'))
            all_final = all.exclude(status__in = ['0','1','7'])
            tot_payy = all_final.aggregate(tot_payy = Sum('face_net_amount'))
            aghd = all.filter(status__in = ['0','1'])
            aghd_pay = aghd.aggregate(tot_pay = Sum('face_net_amount'))
            all_cancel = all.filter(status = '7')
            tot_payy_cancel = all_cancel.aggregate(tot_payy = Sum('face_net_amount'))
            all_final_l1 = all.filter( customer__level = '1').exclude(status__in = ['0','1','7'])
            
            tot_payy_l1 = all_final_l1.aggregate(tot_payy = Sum('face_net_amount'))
            
            all_final_l2 = all.filter( customer__level = '2').exclude(status__in = ['0','1','7'])
            tot_payy_l2 = all_final_l2.aggregate(tot_payy = Sum('face_net_amount'))
            
            all_final_l3 = all.filter( customer__level = '3').exclude(status__in = ['0','1','7'])
            tot_payy_l3 = all_final_l3.aggregate(tot_payy = Sum('face_net_amount'))
            
            all_shop = all.filter(status = '3')
            x1 = 0
            
            for c in all_final:
                x1 += c.supplier_balance
            
            ta = all.filter(status = '2')
            
            x2 = 0
            for c in ta:
                x2 += c.supplier_balance
            
            taa = all.filter(status = '1')
            x3 = 0
            for c in taa:
                x3 += c.supplier_balance
            x4 = 0
            for c in aghd:
                x4 += c.supplier_balance
            
            pa = all.filter(status = '4')
            z = 0
            aa = 0
            sa = 0
            saa = 0
            list_of_ids_takh = []
            list_of_ids_moa = []
            for con in pa:
                po = con.debit_amount()
                if po > 0:
                    qq = po // con.instalment_amount
                    if qq < (con.number_of_instalment// 12) +1:
                        z += 1
                        aa += po
                    else:
                        sa += 1
                        saa += po
           
            l6 = z
            l7 = aa
            l8 = sa
            l9 = saa 
            nok = all.filter(status = '6')
            ee = 0
            for c in nok:
                ee += c.debit_amount()
            
            www = 0
            for c in nok:
                www += c.total_amount_of_instalments
            
            l10 = z + sa +n(nok)
            l11 = ee + saa + aa
            nokk = nok.annotate(pay = Sum('payments__amount')).aggregate(tot_pay = Sum('pay'))
            payy = nokk['tot_pay']
            
            try:
                bedehi = l11 / tot_payments['tot_pay']*100
                bedehi = round(bedehi,2)
            except:
                bedehi = 0
            if payy is  None:
                payy = 0
            
            f = www - payy
            vcc_tot  = z + sa + n(nok)
            vcc_tot_debit = aa + saa + ee
            nok1 = www - payy 
            vcc_q = nok1 + aa + saa
            if tot_payments['tot_pay'] is None:
                tot_payments['tot_pay'] = 1
            nookk = nok1 / tot_payments['tot_pay']
            try:
                nookk = nok1 / tot_payments['tot_pay']*100
                nookk = round(nookk,2)
                
            except:
                nookk = 0
            
            st = []
            for item in supp:
                st.append(item.name)
            
            zzz = n(all_final)
            if zzz == 0:
                zzz = 1
            
            document = MailMerge(os.path.join(INPUT_DIR,"supp_mailmerge.docx"))
            context = {"date" : str(persian_numbers_converter(jalali.Gregorian(date.today()).persian_string('{}/{}/{}'))) ,
              "supplier_name" : ','.join(st) ,
              "tashkil_count" : str(p(n(all))), 
              "tashkil_net_amount" : str(p(tot_pay['tot_pay'])),
              "nahai_count" : str(p(n(all_final))),
              "nahai_net_amount" : str(p(tot_payy['tot_payy'])),
              "aghd_count" : str(p(n(aghd))),
              "aghd_net_amount" : str(p(aghd_pay['tot_pay'])),
              "enseraf_count" : str(p(n(all_cancel))) ,
              "enseraf_net_amount" : str(p(tot_payy_cancel['tot_payy'])),
              "level1_count" : str(p(n(all_final_l1))),
              "level1_net_amount" : str(p(tot_payy_l1['tot_payy'])),
              "level2_count" : str(p(n(all_final_l2))),
              "level2_net_amount" : str(p(tot_payy_l2['tot_payy'])),
              "level3_count" : str(p(n(all_final_l3))),
              "level3_net_amount" : str(p(tot_payy_l3['tot_payy'])),
              "nahai_suppbalance" : str(p(x1)),
              "taid_suppbalance" : str(p(x2)),
              "aghd_suppbalance" : str(p(x4)),
              "vcc_count" : str(p(l10)),
              "vcc_debt" : str(p(l11)),
              "vcc_count_rate" : str(p(round(l10/zzz,3)*100)),
              "vcc_debt_rate" : str(p(bedehi)),
              "nokol_count" : str(p(n(nok))),
              "nokol_total_debt" : str(p(nok1)),
              "nokol_total_debt_rate" : str(p(nookk)), 
            }
            document.merge(**context)
            destination = os.path.join(INPUT_DIR,"supp_mailmerge_out.docx")
            document.write(destination)
            document.close()
            pdf_dest = os.path.join(OUTPUT_DIR,"supp")
            os.system ("libreoffice --headless --convert-to pdf --outdir "+pdf_dest + " "+ destination)
            pdf = os.path.join(pdf_dest,"supp_mailmerge_out.pdf")
            return FileResponse(open(pdf,'rb'),as_attachment = True)
    else:
        finance_form = vccForm()
        return render(request, 'vccForm.html', { 'vccForm' : finance_form , })

@user_passes_test(lambda u : u.is_superuser)
def Cleared(request):
    if request.method == 'POST':
        financing_form = vccForm(request.POST)
        if financing_form.is_valid():
            btn_status = financing_form.cleaned_data['btn']
            if btn_status == True :
                supp = supplier.objects.all()
            else:
                supp = financing_form.cleaned_data["supp"]
            all = contract.objects.filter(supplier__in = supp)
            all_cl_fi = all.exclude(status = '7')

            am_cl_fi = all_cl_fi.aggregate(tot_pay = Sum('face_net_amount'))
            cl = all_cl_fi.filter(status = '5').annotate(pay = Sum('payments__amount'))
            am_cl = cl.aggregate(tot_pay = Sum('face_net_amount'))
            r = 0
            for con in cl:
                r += con.loan_amount
            
            x = 0
            for c in cl:
                if c.pay is not None:
                    x += c.pay
            
            z = 0
            for c in cl:
                z += c.total_amount_of_instalments
            
            profit = z - r
            st = []
            for item in supp:
                st.append(item.name)
            
            sq = []
            for item in supp:
                sq.append(item.coffer.name)
            sq  = set(sq)
            sqq = []
            for item in supp:
                sqq.append(item.issuer.name)
            sqq  = set(sqq)
            zz = n(all_cl_fi)
            if zz == 0:
                zz = 1
            po = round(n(cl)/zz,3)*100
            document = MailMerge(os.path.join(INPUT_DIR,"cleared_mailmerge.docx"))
            context = {"date" : str(persian_numbers_converter(jalali.Gregorian(date.today()).persian_string('{}/{}/{}'))) ,
              "supplier_name" :  ','.join(st),
              "coffer_name" : ','.join(sq), 
              "issuer_name" : ','.join(sqq),
              "activetas_count" : str(p(n(all_cl_fi))),
              "activetas_net_amount" : str(p(am_cl_fi['tot_pay'])),
              "tasvieh_count" : str(p(n(cl))) ,
              "tasvieh_net_amount" : str(p(am_cl['tot_pay'])),
              "tasvieh_rate" : str(p(po)),
              "tasvieh_loan_amount" : str(p(r)),
              "tasvieh_total_insts" : str(p(z)),
              "tasvieh_gain" : str(p(z - r)),
            }
            document.merge(**context)
            destination = os.path.join(INPUT_DIR,"cleared_mailmerge_out.docx")
            document.write(destination)
            document.close()
            pdf_dest = os.path.join(OUTPUT_DIR,"cleared")
            os.system ("libreoffice --headless --convert-to pdf --outdir "+pdf_dest + " "+ destination)
            pdf =   os.path.join(pdf_dest,"cleared_mailmerge_out.pdf")

            return FileResponse(open(pdf,'rb'),as_attachment = True)
    else:
        finance_form = vccForm()
        return render(request, 'vccForm.html', { 'vccForm' : finance_form , })

def income(request):
    if request.method == 'POST':
        financing_form = vccForm(request.POST)
        if financing_form.is_valid():
            btn_status = financing_form.cleaned_data['btn']
            if btn_status == True:
                supp = supplier.objects.all()
            else:
                supp = financing_form.cleaned_data["supp"]
            all = contract.objects.filter(supplier__in = supp )
            all_fi = all.exclude(status__in = ['0','1','7'])
            tot_pay_fi = all_fi.aggregate(tot_payy = Sum('face_net_amount'))
            
            all_cl = all.filter(status = '5').annotate(pay = Sum('payments__amount'))
            all_cl_pay = all.filter(status = '5').annotate(pay = Sum('payments__amount')).aggregate(tot_pay = Sum('pay'))
            x = 0
            for c in all_cl:
                x += c.loan_amount
            baz = 0
            for c in all_fi:
                baz += c.investor_gain
            com = 0
            for c in all_fi:
                com += c.company_gain
            ins = 0
            for c in all_fi:
                ins += c.total_amount_of_instalments
            war = 0
            for c in all_fi:
                war += c.warranty_gain_share
            pom = 0
            for c in all_fi:
                pom += c.pure_company_gain
            tom = 0
            for c in all_fi:
                tom += c.total_company_gain
            st = []
            for item in supp:
                st.append(item.name)
            
            sq = []
            for item in supp:
                sq.append(item.coffer.name)
            sq  = set(sq)
            sqq = []
            for item in supp:
                sqq.append(item.issuer.name)
            sqq  = set(sqq)
            document = MailMerge(os.path.join(INPUT_DIR,"income_mailmerge.docx"))
            context = {"date" : str(persian_numbers_converter(jalali.Gregorian(date.today()).persian_string('{}/{}/{}'))) ,
              "supplier_name" :  ','.join(st),
              "coffer_name" : ','.join(sq), 
              "issuer_name" : ','.join(sqq),
              "nahai_count" : str(p(n(all_fi))),
              "nahai_net_amount" : str(p(tot_pay_fi['tot_payy'])),
              "nahai_loan_amount" : str(p(x)) ,
              "tasvieh_total_insts" : str(p(ins)),
              "investor_gain" : str(p(baz)),
              "company_gain" : str(p(com)),
              "warranty_gain_share" : str(p(war)),
              "pure_company_gain" : str(p(pom)),
              "total_company_gain" : str(p(tom)),
            }
            document.merge(**context)
            destination = os.path.join(INPUT_DIR,"income_mailmerge_out.docx")
            document.write(destination)
            document.close()
            pdf_dest = os.path.join(OUTPUT_DIR,"income")
            os.system ("libreoffice --headless --convert-to pdf --outdir "+pdf_dest + " "+ destination)
            pdf =os.path.join(pdf_dest,"income_mailmerge_out.pdf")
            return FileResponse(open(pdf,'rb'),as_attachment = True)
    else:
        finance_form = vccForm()
        return render(request, 'vccForm.html', { 'vccForm' : finance_form , })