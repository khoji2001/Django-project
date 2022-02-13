from .models import contract,month_summ, vcc_free
from OTP.views import send_sms
from OTP.models import select_MessageText
from django.core.mail import EmailMessage
import jdatetime
from extensions import jalali
from OTP.models import Emails
from extensions.utill import persian_numbers_converter
from django.utils import timezone
import datetime
from supplier.models import coffer, supplier
import string
import xlsxwriter
from customer.models import customer
import os 
import openpyxl
from django.conf import settings
from openpyxl.styles import PatternFill

MONTH_SUMM_EXCEL = {
	'supp_name' : 'نام فروشگاه' ,
    'total_net_amount' : 'مبلغ کلی فروش ماهانه' ,
	'signed_num' : 'تعداد کل قراردادها' ,
    'pending_net_amount' : 'مبلغ در حال عقد قرارداد',
    'pending_num' : 'تعداد در حال عقد قرارداد',
    'other_net_amount' : 'مبلغ سایر وضعیت قرارداد',
    'other_num' : 'تعداد سایر وضعیت قرارداد',
    'total_investor_gain' : 'کارمزد بازاریاب' ,
	'total_company_gain' : 'کارمزد شرکت' ,
	'total_warranty_gain_share' : 'درآمد شرکت از ضمانت نامه' ,
	'total_gain' : 'درآمد کلی شرکت' ,
	'signed_num_1' : 'تعداد سطح ۱' ,
	'total_net_amount_1' : 'مبلغ سطح ۱' ,
	'signed_num_2' : 'تعداد سطح ۲' ,
	'total_net_amount_2' : 'مبلغ سطح ۲' ,
	'signed_num_3' : 'تعداد سطح ۳' ,
	'total_net_amount_3' : 'مبلغ سطح ۳' ,
}

PRE_GENDER = {
	'm' : 'حضرتعالی' , 
	'f' : 'سرکارعالی'
}

def due_contracts():
    due_date = persian_numbers_converter(jdatetime.date.fromgregorian(date = timezone.now().date() + datetime.timedelta(days = 1),locale = "fa_IR").strftime("%a %d %b"))
    due_contracts = contract.due_contracts.all()
    count = len(due_contracts)
    result = ""
    for c in due_contracts:
        try:
            if c.number_of_pays() == 0:
                is_first = "اول "
            else:
                is_first = ''

            if c.status in ['2','3']:
                c.status = '4'
                c.save()
            if c.instalment_amount + c.debit_amount() > 0:
                amount = persian_numbers_converter(c.instalment_amount + c.debit_amount(),'price')
                text = select_MessageText().notify_instalment.format(is_first,due_date,amount,c.vcc_number)
                send_sms(c.customer.user.mobile_number , text)
            else:
              
                text = select_MessageText().good_payer.format(due_date , PRE_GENDER[c.customer.user.gender])
                send_sms(c.customer.user.mobile_number , text)
        except Exception as e:
            result += str(e) + '\n'
    send_sms("09374788634" , "پیامک یادآوری اقساط برای {0}نفر ارسال شد".format(str(count)))

def month_summary():
    today = datetime.date.today()
    year,month,day = jalali.Gregorian(today).persian_tuple()
    if day == 1:
        year,month,day = jalali.Gregorian(today - datetime.timedelta(days=1)).persian_tuple()
        if month < 7: #first half of year
            delta_2 = datetime.timedelta(days=31)
        elif month == 12: #esfand month
            if year % 4 == 3: #kabise year
                delta_2 = datetime.timedelta(days=30)
            else: #not kabise
                delta_2 = datetime.timedelta(days=29)
        else: #second half of year and not esfand
            delta_2 = datetime.timedelta(days=30)
        start = today - delta_2
        stop = today - datetime.timedelta(days=1)
        supps = supplier.objects.all()
        month_name = jdatetime.date.fromgregorian(date = today - datetime.timedelta(days=1),locale = "fa_IR").strftime("%b")
        workbook = xlsxwriter.Workbook(os.path.join(settings.BASE_DIR,"contract_docs/" + month_name + ".xlsx"))
        cell_format = workbook.add_format({'num_format': '#,###'})
        cell_format.set_align('center')
        cell_format.set_align('vcenter')
        cell_format.set_font_name('B Nazanin')
        worksheet = workbook.add_worksheet()
        worksheet.right_to_left()
        row = 0
        col = 0
        worksheet.set_column(1 , 20 , 20 , cell_format)
        for data in MONTH_SUMM_EXCEL.values():
            worksheet.write(row, col , data)
            col += 1
        row += 1
        col = 0
        for s in supps:
            col = 0
            summ = month_summ(s , start,stop)
            summ_dict = vars(summ)
            if summ_dict.get('signed_num') == 0:
                continue
            for key in MONTH_SUMM_EXCEL:
                data = summ_dict.get(key)
                worksheet.write(row , col , data)
                col += 1
            row += 1
        col = 0
        worksheet.write(row,col,'مجموع')
        col += 1
        sum_formula = '=SUM({0}{1}:{0}{2})'
        alpha = list(string.ascii_uppercase)
        for i in range( 1 , len(MONTH_SUMM_EXCEL)):
            val = alpha[i]
            worksheet.write(row,col,sum_formula.format(val,'2',str(row)))
            col += 1
        workbook.close()
        email_dict = {
                'm_n' : month_name,
                }
        
        emaill  = Emails.objects.get(email_type = '2')
        too = [x.strip() for x in emaill.TO.split(',')]
        email = EmailMessage(subject = "ایمیل فروش ماهیانه",body = emaill.ET ,from_email = 'admin@test.com', to = too ,headers= {'Content_Type' :'text/plain'})
        email.attach_file(os.path.join(settings.BASE_DIR,"contract_docs/" + month_name + ".xlsx"))
        email.send()
        

def is_surety():
    last_month = datetime.date.today() - datetime.timedelta(days=30)
    expiredـsurety = customer.objects.filter(surety_date__lt=last_month).filter(surety_permission=True)
    for expired in expiredـsurety:
        expired.surety_permission = False
        expired.save()

def contract_permission_expired():
    Del_date = datetime.date.today() - datetime.timedelta(days=14)
    contr = contract.objects.filter(status = '1').filter(contract_permission_date__lt = Del_date)
    sms_text = select_MessageText().contract_permission_expire
    for item in contr:
        item.status = '7'
        item.save()
        send_sms(item.customer.user.mobile_number, sms_text)
def free_vcc_coffer():
    redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
    coff = coffer.objects.all()
    ex = openpyxl.load_workbook(os.path.join(settings.BASE_DIR,'contract_docs/input/free_vcc.xlsx'))
    sheet = ex['Sheet1']
    sheet.column_dimensions['A'].width=40
    for i,c in enumerate(coff):
        count = vcc_free.objects.filter(coffer = c).count()
        sheet.cell(row=i+2, column=2).value = count
        sheet.cell(row=i+2, column=1).value = c.name
        if count < 10:
            sheet.cell(row=i+2, column=2).fill = redFill
    ex.save(os.path.join(settings.BASE_DIR,'contract_docs/free_vcc_s.xlsx'))
    subject_text = 'اطلاعات کارتهای صندوق'
    email_text = 'باسلام و احترام اطلاعات کارت های هر صندوق به پیوست تقدیم می گردد.'
    email = EmailMessage(subject = subject_text,body = email_text ,from_email = 'admin@test.com', to = ['ited.test@gmail.com',] ,headers= {'Content_Type' :'text/plain'})
    email.attach_file(os.path.join(settings.BASE_DIR,"contract_docs/free_vcc_s.xlsx"))
    email.send()

